#!/usr/bin/env python3
"""
Convert official USDA 2023 RUCC Excel file to JSON format.
"""

import json
import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('rucc_2023_official.xlsx')
ws = wb.active

# Find the header row
headers = []
data_start_row = None

for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
    values = [cell.value for cell in row]
    # Look for FIPS column
    if 'FIPS' in [str(v).upper() if v else '' for v in values]:
        headers = values
        data_start_row = row_idx + 1
        break

if not headers:
    print("ERROR: Could not find header row with FIPS column")
    exit(1)

# Find column indices
fips_col = None
rucc_col = None
state_col = None
county_col = None
pop_col = None

for idx, header in enumerate(headers):
    if header is None:
        continue
    header_upper = str(header).upper()
    if 'FIPS' in header_upper:
        fips_col = idx
    elif 'RUCC' in header_upper and '2023' in header_upper:
        rucc_col = idx
    elif 'STATE' in header_upper:
        state_col = idx
    elif 'COUNTY' in header_upper and 'NAME' in header_upper:
        county_col = idx
    elif 'POPULATION' in header_upper and '2020' in header_upper:
        pop_col = idx

print(f"Found columns - FIPS: {fips_col}, RUCC: {rucc_col}, State: {state_col}, County: {county_col}")

if fips_col is None or rucc_col is None:
    print("ERROR: Could not find required columns")
    exit(1)

# Parse data
county_classifications = {}
rucc_counts = {}

for row in ws.iter_rows(min_row=data_start_row):
    fips = row[fips_col].value
    rucc = row[rucc_col].value

    if not fips or not rucc:
        continue

    # Convert FIPS to 5-digit string
    fips_str = str(fips).strip()
    if len(fips_str) < 5:
        fips_str = fips_str.zfill(5)

    # Convert RUCC to integer
    try:
        rucc_int = int(rucc)
    except (ValueError, TypeError):
        continue

    # Determine category based on RUCC code
    if rucc_int in [1, 2, 3]:
        category = "metro"
    elif rucc_int in [4, 6, 8]:
        category = "nonmetro_adjacent"
    elif rucc_int in [5, 7]:
        category = "nonmetro_nonadjacent"
    elif rucc_int == 9:
        category = "frontier"
    else:
        continue

    county_classifications[fips_str] = {
        "rucc": rucc_int,
        "category": category
    }

    # Count distribution
    rucc_counts[rucc_int] = rucc_counts.get(rucc_int, 0) + 1

# Create complete JSON structure
rucc_data = {
    "metadata": {
        "source": "USDA Economic Research Service",
        "year": 2023,
        "description": "Rural-Urban Continuum Codes (RUCC) - Official Data",
        "download_url": "https://www.ers.usda.gov/data-products/rural-urban-continuum-codes",
        "download_date": "2025-09-30",
        "classification": {
            "1": "Metro - Counties in metro areas of 1 million population or more",
            "2": "Metro - Counties in metro areas of 250,000 to 1 million population",
            "3": "Metro - Counties in metro areas of fewer than 250,000 population",
            "4": "Nonmetro - Urban population of 20,000 or more, adjacent to a metro area",
            "5": "Nonmetro - Urban population of 20,000 or more, not adjacent to a metro area",
            "6": "Nonmetro - Urban population of 2,500 to 19,999, adjacent to a metro area",
            "7": "Nonmetro - Urban population of 2,500 to 19,999, not adjacent to a metro area",
            "8": "Nonmetro - Urban population fewer than 5,000, adjacent to a metro area",
            "9": "Nonmetro - Urban population fewer than 5,000, not adjacent to a metro area"
        },
        "simplified_categories": {
            "metro": [1, 2, 3],
            "nonmetro_adjacent": [4, 6, 8],
            "nonmetro_nonadjacent": [5, 7],
            "frontier": [9]
        },
        "notes": [
            "2023 codes based on 2020 Census and OMB metro area delineations",
            "Urban area population threshold changed from 2,500 to 5,000 in 2023",
            "Some Virginia independent cities combined with surrounding counties"
        ]
    },
    "county_classifications": county_classifications
}

# Save to JSON
output_file = 'Data Preprocessing/county_rucc_2023_official.json'
with open(output_file, 'w') as f:
    json.dump(rucc_data, f, indent=2)

print(f"\n✅ Successfully converted {len(county_classifications)} counties")
print(f"📁 Saved to: {output_file}")

print("\n📊 RUCC Distribution (Official USDA 2023 Data):")
metro_total = sum(rucc_counts.get(i, 0) for i in [1, 2, 3])
nonmetro_total = sum(rucc_counts.get(i, 0) for i in [4, 5, 6, 7, 8, 9])

for rucc_code in sorted(rucc_counts.keys()):
    count = rucc_counts[rucc_code]
    pct = (count / len(county_classifications)) * 100
    print(f"  RUCC {rucc_code}: {count:4d} counties ({pct:5.1f}%)")

print(f"\nTotals:")
print(f"  Metro (1-3):    {metro_total:4d} counties")
print(f"  Nonmetro (4-9): {nonmetro_total:4d} counties")
print(f"  Total:          {len(county_classifications):4d} counties")
